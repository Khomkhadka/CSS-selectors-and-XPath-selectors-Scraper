# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment
# from openpyxl.utils import get_column_letter
# from itemadapter import ItemAdapter
# from datetime import datetime
# import os

# class CleanDataPipeline:
#     """Cleans whitespace from every scraped field."""
#     def process_item(self, item, spider=None):
#         adapter = ItemAdapter(item)
#         for field in ["title","salary","contract_type","job_type","location"]:
#             val = adapter.get(field, "")
#             if val:
#                 adapter[field] = " ".join(val.split()) or "Not specified"
#         return item

# class ExcelExportPipeline:
#     """Saves all jobs to a formatted Excel file."""

#     def open_spider(self, spider=None):
#         # Called ONCE when spider starts — create Excel file
#         self.wb = openpyxl.Workbook()
#         self.ws = self.wb.active
#         self.ws.title = "Data Analyst Jobs"
#         self.count = 0
#         headers = ["Detail URL","Title","Salary",
#                    "Contract Type","Job Type","Location"]
#         widths  = [50, 30, 25, 18, 15, 20]
#         fill = PatternFill("solid", fgColor="1F4E79")
#         font = Font(bold=True, color="FFFFFF", size=11)
#         for i, (h, w) in enumerate(zip(headers, widths), 1):
#             c = self.ws.cell(row=1, column=i, value=h)
#             c.fill = fill; c.font = font
#             c.alignment = Alignment(horizontal="center")
#             self.ws.column_dimensions[get_column_letter(i)].width = w
#         self.ws.freeze_panes = "A2"

#     def process_item(self, item, spider=None):
#         # Called for EACH job — add one row
#         self.count += 1
#         r = self.count + 1
#         adapter = ItemAdapter(item)
#         fill = PatternFill("solid",
#             fgColor="D6E4F0" if self.count % 2 == 0 else "FFFFFF")
#         vals = [adapter.get("detail_url","N/A"),
#                 adapter.get("title","N/A"),
#                 adapter.get("salary","Not specified"),
#                 adapter.get("contract_type","Not specified"),
#                 adapter.get("job_type","Not specified"),
#                 adapter.get("location","Not specified")]
#         for i, v in enumerate(vals, 1):
#             c = self.ws.cell(row=r, column=i, value=v)
#             c.fill = fill
#             if i == 1 and v.startswith("http"):
#                 c.hyperlink = v
#                 c.font = Font(color="0563C1", underline="single")
#         return item

#     def close_spider(self, spider=None):
#         # Called ONCE when spider finishes — save file
#         ts = datetime.now().strftime("%Y%m%d_%H%M%S")
#         name = f"reed_jobs_{self.count}_records_{ts}.xlsx"
#         self.wb.save(name)
#         spider.logger.info(f"✅ Saved {self.count} jobs to {name}")


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from itemadapter import ItemAdapter
from datetime import datetime


class CleanDataPipeline:
    """Cleans whitespace from every scraped field."""

    def process_item(self, item, spider):
        adapter = ItemAdapter(item)

        for field in ["title", "salary", "contract_type", "job_type", "location"]:
            val = adapter.get(field)

            if val:
                adapter[field] = " ".join(str(val).split())

            if not adapter.get(field):
                adapter[field] = "Not specified"

        return item


class ExcelExportPipeline:
    """Saves all jobs to a formatted Excel file."""

    def open_spider(self, spider):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Data Analyst Jobs"

        self.count = 0

        headers = [
            "Detail URL",
            "Title",
            "Salary",
            "Contract Type",
            "Job Type",
            "Location"
        ]

        widths = [50, 30, 25, 18, 15, 20]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = self.ws.cell(row=1, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            self.ws.column_dimensions[get_column_letter(i)].width = w

        self.ws.freeze_panes = "A2"

    def process_item(self, item, spider):
        self.count += 1
        row = self.count + 1

        adapter = ItemAdapter(item)

        values = [
            adapter.get("detail_url", "N/A"),
            adapter.get("title", "Not specified"),
            adapter.get("salary", "Not specified"),
            adapter.get("contract_type", "Not specified"),
            adapter.get("job_type", "Not specified"),
            adapter.get("location", "Not specified"),
        ]

        row_fill = PatternFill(
            "solid",
            fgColor="D6E4F0" if self.count % 2 == 0 else "FFFFFF"
        )

        for col, value in enumerate(values, 1):
            cell = self.ws.cell(row=row, column=col, value=value)
            cell.fill = row_fill

            # hyperlink for URL
            if col == 1 and isinstance(value, str) and value.startswith("http"):
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

        return item

    def close_spider(self, spider):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reed_jobs_{self.count}_records_{ts}.xlsx"

        self.wb.save(filename)

        spider.logger.info(
            f"Saved {self.count} jobs to {filename}"
        )